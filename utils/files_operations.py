#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Apr 13 16:18:01 2019

@author: Jenkins
"""

import re
import os
import csv
import json
import smtplib
import pickle
import openpyxl
import time
import datetime

import numpy as np
import pandas as pd
import multiprocessing as mp

from tqdm import tqdm
from urllib import request
from openpyxl import Workbook
#from urllib import request, parse
from openpyxl import load_workbook



"""date"""
def beforeNday(n, timestr=None):
    """Calculate the next day of the given date."""
    if not timestr: timestr = time.strftime("%Y-%m-%d", time.localtime())
    date = datetime.datetime.strptime(timestr, '%Y-%m-%d').date()
    return (date + datetime.timedelta(-n)).strftime('%Y-%m-%d')


def dateDiff(dt1, dt2):
    dt1 = datetime.datetime.strptime(dt1,'%Y-%m-%d')
    dt2 = datetime.datetime.strptime(dt2,'%Y-%m-%d')
    return (dt1 - dt2).days


"""files op"""
def files_sorting(files):
    '''Sort the reading files'''
    lenfiles = len(files)    
    sorted_mat, sorted_files = np.zeros((lenfiles,1)), [0]*lenfiles  
    for i in range(lenfiles):
        sorted_mat[i] = re.sub("\D", "", files[i])    
    sort_ind =  np.argsort(sorted_mat.astype(int),axis=0)
    for i in range(lenfiles):
        sorted_files[i] = files[sort_ind[i][0]] 
    return sorted_files


def files_finding(paths, sort=True):
    '''Files finding'''
    o_files = os.listdir(paths)
    files, len_files = [], len(o_files)
    for i in range(len_files):
        if o_files[i].find('~$') != 0: files.append(o_files[i])
    if sort: return files_sorting(files), len(files)
    else: return files, len(files)


def create_folder(path):
    sp_folder = path.split('/')
    folder = sp_folder[0]
    for i in range(len(sp_folder)-1):
        if not os.path.exists(folder): os.mkdir(folder)
        folder = folder + '/' + sp_folder[i+1]
        
        
def data_normalization(data):
    return (data-data.min(0))/(data.max(0)-data.min(0))   


def iter_split(iter_num, num):
    l = len(iter_num)
    p, q = l%num, l//num
    res = []
    count = 0
    for i in range(num):
        tmp = count+q+ (1 if i<p else 0)
        res.append(iter_num[count:tmp])
        count = tmp 
    return res



"""saving and reading"""
def pd_save(data, path, sheetname='sheet1', ind=False):
    writer = pd.ExcelWriter(path)
    data.to_excel(excel_writer=writer, sheet_name=sheetname, index = ind)
    writer.save()
    # writer.close()



def xlsx_save(data, path, header=False):
    create_folder(path)
    if len(data.shape) == 1: data = data[:, np.newaxis]
    wb = Workbook()    
    ws = wb.active
    if header: ws.append(header)      
    [h, l] = data.shape  # h:rows，l:columns
    for i in range(h):
        row = []
        for j in range(l):
            row.append(data[i,j])
        ws.append(row)
    wb.save(path)
    

def xlsx_sheet_save(data, path, sheetname, ind=0):  # multiple sheets
    create_folder(path)
    wb = openpyxl.load_workbook(path)
    wb.create_sheet(index=ind, title=sheetname)
    ws = wb[sheetname]     
    [h, l] = data.shape  # h:rows，l:columns
    for i in range(h):
        row = []
        for j in range(l):
            row.append(data[i,j])
        ws.append(row)
    wb.save(path)   
    

def csv_write(data, path, header=None):
    with open(path, 'w', newline='', encoding='utf-8-sig') as csvfile:
        spamwriter = csv.writer(csvfile)
        if header:
            spamwriter.writerow(header)
        spamwriter.writerows(data)
               

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
    Returns: None
    """
    # from openpyxl import load_workbook

    # import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def pickle_save(data, path):    
    obj1 = pickle.dumps(data)
    with open(path, 'wb') as f:
        f.write(obj1)
    f.close()
    
    
def pickle_read(path):
    with open(path, 'rb') as df:
        data = pickle.load(df)
    df.close()  
    return data
    


"""alarming"""
def beep():
    print('\a'*7)
 
    
def message_alarm():
    id_code = "tyv9Su9"
    request.urlopen("http://miaotixing.com/trigger?id=%s"%id_code)


def send_message(email='630158324@qq.com'):    
    sender = email
    receivers = [email]
    message = """From:From Person <from@fromdomain.com>
    To: To Person <to@todomain.com>
    Subject: Running Complete !
    
    Running complete! Hurry up checking the results!
    """
    
    smtpObj = smtplib.SMTP('smtp.qq.com')
    smtpObj.login(email,'tynmbjekdhoxbfec')
    smtpObj.sendmail(sender, receivers, message)
    smtpObj.quit()


def tqdm_close():
    try:
        with tqdm(...) as t:
            for i in t:
                pass
    except KeyboardInterrupt:
        t.close()
        raise
    t.close()
    
